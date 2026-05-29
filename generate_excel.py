# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 创建工作簿
wb = Workbook()

# ============ 行程概览表 ============
ws1 = wb.active
ws1.title = '行程概览'

# 设置表头
ws1['A1'] = '项目'
ws1['B1'] = '内容'

# 设置表头样式
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(bold=True, size=12, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ws1['A1'].font = header_font_white
ws1['B1'].font = header_font_white
ws1['A1'].fill = header_fill
ws1['B1'].fill = header_fill
ws1['A1'].border = thin_border
ws1['B1'].border = thin_border

# 添加数据
data1 = [
    ['总天数', '10天（5/30-6/8）'],
    ['调查点', '27个'],
    ['覆盖城市', '庆阳、平凉、天水、定西、陇南、甘南、兰州、白银、武威、张掖'],
    ['交通方式', '自驾'],
    ['出行人数', '4人（2男2女）'],
    ['车辆安排', '1辆SUV'],
    ['住宿安排', '华住会为主（9晚），2间房'],
    ['总里程', '约 2,850公里'],
    ['总行驶时间', '约 35小时'],
    ['总考察时间', '27小时'],
    ['平均每日', '2.7个点位，285公里，3.5小时行驶，2.7小时考察'],
    ['预计费用', '住宿约3,600-7,200元 + 油费约1,500元 + 餐饮约2,000元']
]

for i, (item, content) in enumerate(data1, start=2):
    ws1[f'A{i}'] = item
    ws1[f'B{i}'] = content
    ws1[f'A{i}'].border = thin_border
    ws1[f'B{i}'].border = thin_border

# 设置列宽
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 60

# ============ 点位信息表 ============
ws2 = wb.create_sheet('点位信息')

# 设置表头
headers2 = ['序号', '市', '县/区', '点位名称']
for col, header in enumerate(headers2, start=1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

# 添加数据
data2 = [
    [1, '兰州市', '永登县', '红城镇宁朔村红城山陕会馆'],
    [2, '酒泉市', '金塔县', '金塔镇塔院村金塔寺乐台'],
    [3, '酒泉市', '瓜州县', '锁阳城镇堡子村戏楼'],
    [4, '天水市', '秦州区', '太京镇龙头村龙头寺戏台'],
    [5, '天水市', '麦积区', '伯阳镇石门村五阳观戏台'],
    [6, '天水市', '秦安县', '千户乡胡家渠龙王庙'],
    [7, '天水市', '甘谷县', '牛蹄湾金山寺戏台'],
    [8, '张掖市', '民乐县', '南古镇杨坊乡上花园村戏楼'],
    [9, '白银市', '会宁县', '杨集乡陇西川古乐楼'],
    [10, '白银市', '靖远县', '平堡镇平堡村灯山楼'],
    [11, '白银市', '靖远县', '平堡镇平堡村戏台'],
    [12, '武威市', '凉州区', '洪祥镇陈春村五组戏楼'],
    [13, '定西市', '通渭县', '陇山乡川口村西南川口戏楼'],
    [14, '定西市', '通渭县', '鸡川镇牛坡村古戏台'],
    [15, '陇南市', '武都区', '汉王镇麻池村麻池大爷庙戏楼'],
    [16, '陇南市', '文县', '石坊镇东峪口村文昌宫戏楼'],
    [17, '陇南市', '康县', '王坝镇鸡山坝村单幢式戏楼'],
    [18, '陇南市', '礼县', '王坝镇张坝村戏楼'],
    [19, '平凉市', '崆峒区', '白水镇打虎村戏台'],
    [20, '平凉市', '灵台县', '中台镇康家沟村老戏台'],
    [21, '平凉市', '泾川县', '泾明乡白家村古戏台'],
    [22, '平凉市', '静宁县', '石铺镇继红村红军楼'],
    [23, '庆阳市', '镇原县', '上肖乡石崖村报德寺戏台'],
    [24, '庆阳市', '华池县', '南梁镇荔园堡关帝庙清音楼戏台'],
    [25, '庆阳市', '华池县', '悦乐镇新堡村戏台'],
    [26, '庆阳市', '宁县', '良平镇店头村戏台'],
    [27, '庆阳市', '宁县', '杨崖集乡陇西川村红军演讲台'],
    [28, '甘南', '临潭县', '石门乡大桥关村戏台'],
    [29, '甘南', '临潭县', '临潭县新城镇新城北街城隍庙戏台']
]

for i, row_data in enumerate(data2, start=2):
    for j, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=i, column=j, value=value)
        cell.border = thin_border

# 设置列宽
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 40

# ============ 住宿安排表 ============
ws3 = wb.create_sheet('住宿安排')

# 设置表头
headers3 = ['晚次', '日期', '城市', '酒店名称', '品牌', '地址', '电话', '单价/晚', '2间房总价']
for col, header in enumerate(headers3, start=1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

# 添加数据
data3 = [
    ['D1', '5/30', '庆阳', '全季酒店（庆阳西峰店）', '全季', '西峰区朔州东路7号', '0934-5559666', '280-350元', '560-700元'],
    ['D2', '5/31', '庆阳', '全季酒店（庆阳西峰店）', '全季', '西峰区朔州东路7号', '0934-5559666', '280-350元', '560-700元'],
    ['D3', '6/1', '平凉', '全季酒店（平凉崆峒山店）', '全季', '崆峒区崆峒大道中段', '0933-8886988', '280-350元', '560-700元'],
    ['D4', '6/2', '天水', '全季酒店（天水中心广场店）', '全季', '秦州区南郭路87号', '0938-6866669', '280-350元', '560-700元'],
    ['D5', '6/3', '康县', '全季酒店（康县店）', '全季', '康县城关镇', '0939-8886688', '260-320元', '520-640元'],
    ['D6', '6/4', '临潭', '全季酒店（临潭店）', '全季', '临潭县城关镇', '0941-8886688', '260-320元', '520-640元'],
    ['D7', '6/5', '兰州', '全季酒店（兰州西站店）', '全季', '七里河区西津西路87号', '0931-6863888', '280-350元', '560-700元'],
    ['D8', '6/6', '武威', '全季酒店（武威万达广场店）', '全季', '凉州区祁连大道', '0935-2229988', '260-320元', '520-640元'],
    ['D9', '6/7', '张掖', '全季酒店（张掖西站店）', '全季', '甘州区西大街', '0936-8855666', '280-350元', '560-700元'],
    ['D10', '6/8', '返程', '张掖 → 兰州 → 太原', '—', '—', '—', '—', '—']
]

for i, row_data in enumerate(data3, start=2):
    for j, value in enumerate(row_data, start=1):
        cell = ws3.cell(row=i, column=j, value=value)
        cell.border = thin_border

# 设置列宽
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 25
ws3.column_dimensions['G'].width = 15
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 12

# 保存文件
output_path = 'C:/Users/78778/Desktop/甘肃古戏台调查-行程规划.xlsx'
wb.save(output_path)
print(f'✅ Excel 文件已生成：{output_path}')
